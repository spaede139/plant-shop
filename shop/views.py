from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.db.models import Q
from django.core.mail import EmailMessage
from django.conf import settings
from django.contrib import messages


from rest_framework import viewsets, permissions, generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.decorators import action

from .models import (
    Product, 
    Category, 
    Manufacturer, 
    Cart, 
    CartItem, 
    Order, 
    OrderItem, 
    Profile
)
from .serializers import (
    ProductSerializer, 
    CategorySerializer, 
    ManufacturerSerializer,
    CartSerializer, 
    CartItemSerializer, 
    UserSerializer, 
    ProfileSerializer, 
    RegisterSerializer
)


import io
import openpyxl



def generate_excel_receipt(order):
    """Генерация Excel чека"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Чек_{order.id}"
    
    ws['A1'] = "ЧЕК ПОКУПКИ"
    ws['A2'] = f"Заказ №{order.id}"
    ws['A3'] = f"Дата: {order.created_at.strftime('%d.%m.%Y %H:%M')}"
    ws['A4'] = f"Покупатель: {order.user.username}"
    ws['A5'] = f"Email: {order.user.email}"
    ws['A6'] = f"Адрес: {order.address}"
    ws['A7'] = f"Телефон: {order.phone}"
    ws['A9'] = "Товары:"
    
    ws['A10'] = "№"
    ws['B10'] = "Товар"
    ws['C10'] = "Кол-во"
    ws['D10'] = "Цена"
    ws['E10'] = "Сумма"

    row = 11
    for idx, item in enumerate(order.items.all(), 1):
        ws[f'A{row}'] = idx
        ws[f'B{row}'] = item.product.name
        ws[f'C{row}'] = item.quantity
        ws[f'D{row}'] = f"{item.price} руб"
        ws[f'E{row}'] = f"{item.quantity * item.price} руб"
        row += 1

    ws[f'D{row+1}'] = "ИТОГО:"
    ws[f'E{row+1}'] = f"{order.total_price} руб"
    ws[f'A{row+3}'] = "Спасибо за покупку!"

    file = io.BytesIO()
    wb.save(file)
    file.seek(0)
    return file




def home(request):
    """Главная страница"""
    popular_products = Product.objects.all().order_by('-id')[:6]
    categories = Category.objects.all()
    return render(request, 'shop/index.html', {
        'popular_products': popular_products,
        'categories': categories,
    })


def about_shop(request):
    """Страница о магазине"""
    return HttpResponse("""
    <h1>О нашем магазине</h1>
    <p>Тема Магазин товаров для ухода за комнатными растениями (лейки, удобрения).</p>
    <a href="/">На главную</a>
    """)


def about_author(request):
    """Страница об авторе"""
    return HttpResponse("""
    <h1>Об авторе</h1>
    <p>Лабораторную сделал Рубцов Петр Группа 89 ТП</p>
    <a href="/">На главную</a>
    """)


def catalog(request):
    """Страница каталога с фильтрацией"""
    products = Product.objects.all()
    
    category = request.GET.get('category')
    manufacturer = request.GET.get('manufacturer')
    search = request.GET.get('search')
    
    if category:
        products = products.filter(category_id=category)
    if manufacturer:
        products = products.filter(manufacturer_id=manufacturer)
    if search:
        products = products.filter(Q(name__icontains=search) | Q(description__icontains=search))
    
    return render(request, 'shop/catalog.html', {
        'products': products,
        'categories': Category.objects.all(),
        'manufacturers': Manufacturer.objects.all(),
        'selected_category': category,
        'selected_manufacturer': manufacturer,
        'search_query': search,
    })


def product_detail(request, pk):
    """Детальная страница товара"""
    return render(request, 'shop/product_detail.html', {
        'product': get_object_or_404(Product, id=pk)
    })


@login_required
def profile_view(request):
    """Личный кабинет"""
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'shop/profile.html', {
        'orders': orders,
    })




@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    cart, _ = Cart.objects.get_or_create(user=request.user)
    item, created = CartItem.objects.get_or_create(
        cart=cart, 
        product=product,
        defaults={'quantity': 1}
    )

    if not created and item.quantity < product.stock:
        item.quantity += 1
        item.save()

    return redirect('cart_detail')


@login_required
def update_cart_item(request, item_id):
    item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    
    if request.method == 'POST':
        qty = int(request.POST.get('quantity', 1))
        if qty > 0 and qty <= item.product.stock:
            item.quantity = qty
            item.save()
        elif qty <= 0:
            item.delete()
    
    return redirect('cart_detail')


@login_required
def remove_from_cart(request, item_id):
    get_object_or_404(CartItem, id=item_id, cart__user=request.user).delete()
    return redirect('cart_detail')


@login_required
def cart_detail(request):
    cart, _ = Cart.objects.get_or_create(user=request.user)
    return render(request, 'shop/cart_detail.html', {
        'cart': cart,
        'items': cart.cartitem_set.all()
    })


@login_required
def checkout(request):
    cart, _ = Cart.objects.get_or_create(user=request.user)
    items = cart.cartitem_set.all()
    
    if not items:
        messages.error(request, 'Ваша корзина пуста!')
        return redirect('cart_detail')
    
    if request.method == 'POST':
        address = request.POST.get('address')
        phone = request.POST.get('phone')
        
        if not address or not phone:
            messages.error(request, 'Пожалуйста, заполните все поля!')
            return render(request, 'shop/checkout.html', {'items': items, 'cart': cart})

        order = Order.objects.create(
            user=request.user,
            address=address,
            phone=phone,
            total_price=cart.total_price()
        )
        
        for item in items:
            OrderItem.objects.create(
                order=order,
                product=item.product,
                quantity=item.quantity,
                price=item.product.price
            )
            item.product.stock -= item.quantity
            item.product.save()
        
        excel_file = generate_excel_receipt(order)

      
        try:
            email = EmailMessage(
                subject=f'Чек заказа #{order.id} в plantcare shop',
                body=f'''
Здравствуйте, {request.user.username}!

Ваш заказ #{order.id} успешно оформлен.

Детали заказа:
━━━━━━━━━━━━━━━━━━━━━━━
Сумма заказа: {order.total_price} руб.
Адрес доставки: {address}
Телефон: {phone}
━━━━━━━━━━━━━━━━━━━━━━━

Чек во вложении в формате Excel.

Спасибо за покупку в plantcare shop!
                ''',
                from_email='djangoshoppt@gmail.com',
                to=[request.user.email],
            )
            
            email.attach('receipt.xlsx', excel_file.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            email.send(fail_silently=False)
            
            messages.success(request, f'Заказ #{order.id} оформлен! Чек отправлен на {request.user.email}')
            
        except Exception as e:
            messages.warning(request, f'Заказ оформлен, но письмо не отправлено: {e}')

        items.delete()
        
        return redirect('catalog')
    
    return render(request, 'shop/checkout.html', {'items': items, 'cart': cart})



class IsAdminOrReadOnly(permissions.BasePermission):
    """Разрешает GET всем, остальные методы только админам"""
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.is_authenticated and (request.user.profile.is_admin() or request.user.is_superuser)


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]


class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [IsAdminOrReadOnly]


class CartViewSet(viewsets.ModelViewSet):
    queryset = Cart.objects.all()
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return Cart.objects.filter(user=self.request.user)


class CartItemViewSet(viewsets.ModelViewSet):
    queryset = CartItem.objects.all()
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return CartItem.objects.filter(cart__user=self.request.user)



class RegisterView(APIView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response({
                'user': UserSerializer(user).data,
                'message': 'Регистрация успешна!'
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LoginView(APIView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return Response({
                'user': UserSerializer(user).data,
                'profile': ProfileSerializer(user.profile).data,
                'message': 'Вход выполнен успешно!'
            })
        return Response({'error': 'Неверные учетные данные'}, status=status.HTTP_401_UNAUTHORIZED)


class LogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def post(self, request):
        logout(request)
        return Response({'message': 'Вы вышли из системы'})


class MeView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        serializer = ProfileSerializer(request.user.profile)
        return Response(serializer.data)
    
    def patch(self, request):
        serializer = ProfileSerializer(request.user.profile, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)